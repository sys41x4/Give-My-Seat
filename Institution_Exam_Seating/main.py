# Developed by Arijit Bhowmick [sys41x4]
# Lisence Type : GNU AGPLv3 [GNU Affero General Public License v3.0]
# https://choosealicense.com/licenses/agpl-3.0/

# https://github.com/Arijit-Bhowmick/Give-My-Seat
# https://github.com/sys41x4/Give-My-Seat


import os
import openpyxl
from pandas import read_excel as pd_read_excel
from numpy import array as np_array
import json
import random

class GiveMySeat:
	def print_indv_branch_table(table):
		global row, col
		global current_block, current_room
		# print individual branch location in matrix
		

		for branch in list(branch_frequency.keys()):
			table_matrix=[]
			for i in range(row):table_matrix+=[['']*col]
			for i in range(len(table)):
				for j in range(len(table[i])):
					if table[i][j]==branch:
						table_matrix[i][j]=table[i][j]

			GiveMySeat.print_table(table_matrix, f'{branch} Seating [{current_block}-Block {current_room}]')



	def print_table(table, matrix_header='MATRIX:'):
					
	    longest_cols = [
	    	(max([len(str(row[i])) for row in table]) + 1)
	    	for i in range(len(table[0]))
	    ]

	    row_format = "".join(["{:>" + str(longest_col) + "}" for longest_col in longest_cols])

	    print(f'[{matrix_header}]\n')
	    for row_num in range(len(table)):
	    	print(row_format.format(*table[row_num]))
	    print()

	def get_rand_branch(table_matrix, prev_value, matrix_row=0, matrix_col=0, upper_value=0):
		global branch_frequency
		global exception_handler

		# Generate modified unique branch values
		# if branch value in branch_frequency==0
		# then it will be excluded in exc_branch list
		exc_branch=[]
		for i in range(len(branch_frequency.values())):
			if list(branch_frequency.values())[i]>0:
				exc_branch+=[list(branch_frequency.keys())[i]]

		
		if exc_branch==[]:
			return '\n'
		###
		else:
			if upper_value!=0:
				try:

					if matrix_row>0 and matrix_col==0:
						if len(exc_branch)==1 and exc_branch[0]==upper_value:
							#return upper_value
							return ''
						exc_branch.remove(upper_value)
					# Except First Entry
					# when .remove() will not give Value Error
					elif matrix_row>0 and matrix_col>0:

						if len(exc_branch)==1 and exc_branch[0]==upper_value and exc_branch[0]==prev_value:
							return ''

						else:
							

							if len(exc_branch)==1 and exc_branch[0]==upper_value:
								return ''
							else:
								if upper_value!='':
									exc_branch.remove(upper_value)

							if len(exc_branch)==1 and exc_branch[0]==prev_value:
								return ''
							else:
								if prev_value!='':
									exc_branch.remove(prev_value)
				except ValueError:
					exception_handler=1

			elif upper_value==0:
				try:
					# Except First Entry
					# when .remove() will not give Value Error
					if len(exc_branch)==1 and exc_branch[0]==prev_value:
							#return prev_value
							return ''
					else:
						if prev_value!='':
							exc_branch.remove(prev_value)
				except ValueError:
					exception_handler=1

			random_branch = random.choice(exc_branch)

			return random_branch

	def remove_confirmed_rollno(seating_json_data):

		global student_data

		# Roll Number List
		rollnum_list = list(student_data['std_rollnum'].values())
		rollnum_index_list = list(student_data['std_rollnum'].keys())

		for roll in list(seating_json_data.keys()):
			for column_header in list(student_data.keys()):
				del student_data[column_header][rollnum_index_list[rollnum_list.index(roll)]]

		print(student_data)

	def remove_filledup_room():
		global current_room
		global room_data

		room_index_list = list(room_data['room_name'].keys())
		room_list = list(room_data['room_name'].values())

		for column_header in list(room_data.keys()):
			del room_data[column_header][room_index_list[room_list.index(current_room)]]


	def json_data_dump(dump_json_data, excel_data_type):

		if excel_data_type=='student_data':
			with open(f"{excel_data_type}.json", 'w', encoding = 'utf-8') as json_file:json.dump(dump_json_data, json_file)
		elif excel_data_type=='room_data':
			with open(f"{excel_data_type}.json", 'w', encoding = 'utf-8') as json_file:json.dump(dump_json_data, json_file)
		elif excel_data_type=='seating_data':
			with open(f"{excel_data_type}.json", 'w', encoding = 'utf-8') as json_file:json.dump(dump_json_data, json_file)


	def arrange_seating(table, student_data, value_to_show):
		global row, col
		global seating_json_data

		table_matrix=[]
		for i in range(row):table_matrix+=[['']*col]


		branch_list = list(student_data['std_branch'].values())

		# Student Name List
		std_name_list = list(student_data['std_name'].values())

		# Student Roll Number List
		rollnum_list = list(student_data['std_rollnum'].values())

		# Student Aadhar Card List
		aadhar_list = list(student_data['std_aadhar'].values())

		# Confirmed Seating Data
		seating_json_data={}

		for row_num in range(len(table)):
			
			for col_num in range(len(table[row_num])):
				if table[row_num][col_num]=='':
					continue

				else:
					branch_index = branch_list.index(table[row_num][col_num])

					branch_name = branch_list[branch_index] # branch Name
					roll_num = rollnum_list[branch_index] # Usually Enrollment Number
					aadhar_num = aadhar_list[branch_index] # Aadhar Number
					student_name = std_name_list[branch_index] # Student Name

					table_matrix[row_num][col_num]=roll_num

					# Creating JSON Data for the seating
					seating_json_data.update({roll_num: {"std_name": student_name,"branch_name": branch_name,"roll_num": roll_num,"aadhar_num": aadhar_num,"block_num": current_block,"room_num": current_room,"row": row_num,"column": col_num }})

					branch_list[branch_index]='\n'

		all_seating_data.update(seating_json_data)
		print()

		# When 1 Room has been successfully booked
		# the students (Based on their Enrollment Number)
		# will be removed from the unassigned list
		# as their seating is confirmed
		GiveMySeat.remove_confirmed_rollno(seating_json_data)

		# Remove the Filleup Room name and its data Row from the json data
		GiveMySeat.remove_filledup_room()

		# Generated Seating data will be dumped in seating_data.json
		GiveMySeat.json_data_dump(all_seating_data, 'seating_data')
		# Generated student data will be dumped in the student_data.json
		GiveMySeat.json_data_dump(student_data, 'student_data')
		# Generated room data will be dumped in the room_data.json
		GiveMySeat.json_data_dump(room_data, 'room_data')


		# On completing 1 room next room would be used

		GiveMySeat.print_table(table_matrix, f'Seating Based on [{current_block}-{current_room}]:')

		# Assign Seating for Remainig Students | Enrollment_No

		GiveMySeat.load_json("from_pre_json", default_json_file_paths)

		# total_student_count
		GiveMySeat.generate_seating(student_data, room_data)




	def generate_matrix_samples(student_data, total_branches):

		

		# branch_frequency is the frequency of
		# branch in the excel sheet

		global branch_frequency
		branch_frequency={}
		for i in (list(set(total_branches))):
			branch_frequency.__setitem__(i, total_branches.count(i))


		## Generate table as matrix
		table_matrix = []
		for i in range(row):table_matrix+=[['']*col]


		exc_branch=list(branch_frequency.keys())

		## if total unique branch are more than column then
		row_finish,matrix_completed=0,0
		matrix_row,matrix_col=0,0
		prev_value=''
		finish=0
		while matrix_completed!=row*col and finish==0:

			if matrix_row==0:
				while row_finish!=col:
					

					table_matrix[matrix_row][matrix_col]=GiveMySeat.get_rand_branch(table_matrix, prev_value)
					prev_value=table_matrix[matrix_row][matrix_col]
					# On adding 1 entry of branch name 
					# in table_matrix, 1 entry/value
					# will be deleted from the branch_frequency dict 
					branch_frequency[prev_value]-=1
					## print(row_finish, col)
					row_finish+=1
					matrix_completed+=1
					matrix_col+=1

				matrix_col=0
				matrix_row+=1
				row_finish=0

			elif matrix_row!=0:
				while row_finish!=col:
					ran_branch = GiveMySeat.get_rand_branch(table_matrix, prev_value, matrix_row, matrix_col, table_matrix[matrix_row-1][matrix_col])

					if ran_branch=='\n':
						finish=1
						break

					table_matrix[matrix_row][matrix_col]=ran_branch
					
					prev_value=ran_branch

					if prev_value!='':
						branch_frequency[prev_value]-=1

					row_finish+=1
					matrix_completed+=1
					matrix_col+=1

				matrix_col=0
				matrix_row+=1
				row_finish=0
		
		# print individual branch location in matrix

		return table_matrix


	def generate_matrix(student_data, sample_count=1):
		# Creating 10 sample matrix (Default Sample_Count: 1)
		# for maximum throughput

		global current_block, current_room
		total_branches = list(student_data['std_branch'].values())

		if len(total_branches)>row*col:
			print(f"Mimimum {len(total_branches)-row*col} students will have to be allocated to another room")

		matrix_samples=[]
		empty_seat_count=[]

		for sample in range(sample_count):
			matrix_samples+=[GiveMySeat.generate_matrix_samples(student_data, total_branches)]


		for sample in matrix_samples:
			empty_seat_count+=[list(np_array(sample).ravel()).count('')]

		max_seating_matrix = matrix_samples[empty_seat_count.index(min(empty_seat_count))]

		GiveMySeat.print_indv_branch_table(max_seating_matrix)

		# Print the formed matrix
		GiveMySeat.print_table(max_seating_matrix, f'Overall Branch Seating in [{current_block}-Block {current_room}:')

		# Print the roll numbers in their provided seating
		GiveMySeat.arrange_seating(max_seating_matrix, student_data, 'std_rollnum')


	def load_json_from_file(json_data_to_use):
		# Reading From JSON Files
		# json format
		# {"student_data_path":"", "room_data_path":"", "seating_data_path":""}

		global student_data
		global room_data
		global all_seating_data
		## Read Student Data from JSON File # Variable Name student_data
		student_data=json.load(open(json_data_to_use["student_data_path"]))
		
		## Read Room Data from JSON File # Variable name room_data
		room_data=json.load(open(json_data_to_use["room_data_path"]))


		## Read Seating Data from JSON File # Variable Name all_seating_data
		all_seating_data=json.load(open(json_data_to_use["seating_data_path"]))

		if all_seating_data=='':
			all_seating_data={}


	def load_json(process_to_use, json_data_to_use):

		if process_to_use=="from_workbook":
			# Generate JSON data from WorkBook
			# and dump it into JSON Files
			# WorkBook settings Format
			# {"workbook_file_path":"", "student_data_sheet_name":"", "room_data_sheet_name":""}

			global student_data
			global room_data
			global all_seating_data

			all_seating_data={}

			student_data=json.loads(pd_read_excel(open(json_data_to_use["workbook_file_path"], 'rb'), sheet_name=json_data_to_use["student_data_sheet_name"]).to_json())
			room_data = json.loads(pd_read_excel(open(json_data_to_use["workbook_file_path"], 'rb'), sheet_name=json_data_to_use["room_data_sheet_name"]).to_json())
			
			GiveMySeat.json_data_dump(student_data, "student_data")
			GiveMySeat.json_data_dump(student_data, "room_data")
			GiveMySeat.json_data_dump(all_seating_data, "seating")

		elif process_to_use=="from_pre_json":
			
			GiveMySeat.load_json_from_file(json_data_to_use)


	def generate_seating(student_json_data, room_json_data):

		global student_data
		global room_data
		global total_student_count # Total Student Number

		total_student_count = len(student_data['std_rollnum'].values())

		if total_student_count==0:
			print("All Students has been assigned to their calculated seating")
			exit()

		# Gather Room Data
		GiveMySeat.gather_room_data(room_data)
		
		# Generate Seating Matrix
		GiveMySeat.generate_matrix(student_data, sample_count)

	def gather_room_data(room_data):
		# Room and its row & Column setup
		global block
		global room_name
		global current_room_count

		global current_block
		global current_room
		global row, col


		row_list = list(room_data['row'].values())
		column_list = list(room_data['column'].values())

		# Current Room data
		current_block = list(room_data['block_name'].values())[0]
		current_room = list(room_data['room_name'].values())[0]
		row = list(room_data['row'].values())[0]
		col = list(room_data['column'].values())[0]

		total_seat_provided = 0
		for i in range(len(row_list)):
			total_seat_provided+=row_list[i]*column_list[i]

		if total_seat_provided<total_student_count:
			print(f"""\nTotal Student Count = {total_student_count}
Total Seats Provided = {total_seat_provided}
Total Student Number is greater than Total Seating
Please add additional rooms with additional seating in Excel Sheet/JSON File\n""")
			exit()

			
			

	def write_to_excel(file_path, sheet_name, matrix, room_name=''):
		col_name_lst = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
		## For xls File
		# # load the excel file
		# load_excel_file = xlrd.open_workbook(file_path)
		 
		# # copy the contents of excel file
		# copy_excel_content = copy(load_excel_file)
		 
		# # open the first sheet
		# w_sheet = copy_excel_content.get_sheet(0)
		 
		# # row number = 0 , column number = 1
		# w_sheet.write(0,1,'Modified !')
		 
		# # save the file
		# wb.save('UserBook.xls')

		## For .xlsx file
		workbook = openpyxl.load_workbook(file_path)

		# The workbook object is then used to add new
		# worksheet via the add_worksheet() method.
		worksheet = workbook.sheetnames[sheet_name]
		 
		# Use the worksheet object to write
		# data via the write() method.
		worksheet.write('A1', 'Hello..')
		worksheet.write('B1', 'Geeks')
		worksheet.write('C1', 'For')
		worksheet.write('D1', 'Geeks')
		 
		# Finally, close the Excel file
		# via the close() method.
		workbook.save('text2.xlsx')

	def main():
		#global row, col
		global sample_count
		global default_json_file_paths

		# Default File Storage Path
		default_json_file_paths={"student_data_path":"student_data.json", "room_data_path":"room_data.json", "seating_data_path":"seating_data.json"}

		sample_count=10
		#row,col=13,4

		#file_path = 'data.xlsx'
		#sheet_name = 'Sheet3'
		#room_detail_sheet = 'allowed_room'
		workbook_data = {"workbook_file_path":"", "student_data_sheet_name":"", "room_data_sheet_name":""}

		#pre_json_data_available='' # Considered as No when entered (Default: NO)
		pre_json_paths={"student_data_path":"", "room_data_path":"", "seating_data_path":""}
		
		try:
			pre_json_data_available = input("Do You have already any pre json data created using this script ? Y/N ==> ")


			if pre_json_data_available.upper()=='Y':
				process_to_use = 'from_pre_json'
				for path_key in list(pre_json_paths.keys())[:-1]:
					ask_path = input(f"\nPlease Enter File_Path of \"{path_key[:-5]}\" => ")
					if os.path.exists(ask_path)==True:
						pre_json_paths.update({path_key:ask_path})
					else:
						return print(f"File Path \"{ask_path}\" is incorrect for \"{path_key[:-5]}\"\nExiting Program !!!")

				ask_path = input(f"Please Enter File_Path of seating_data => ")

				if ask_path=='':
					pre_json_paths.update({"seating_data_path":''})
				elif os.path.exists(ask_path)==True:
					pre_json_paths.update({"seating_data_path":ask_path})
				else:
					return print(f"File Path {ask_path} is incorrect for \"seating_data\"\nExiting Program !!!")

				json_data_to_use = pre_json_paths
			else:
				process_to_use = 'from_workbook'
				ask = input(f"Please Enter File_Path of {list(workbook_data.keys())[0][:-5]} => ")
				if os.path.exists(ask)==True:
					workbook_data.update({"workbook_file_path":ask})
				else:
					return print(f"File Path {ask} is incorrect\nExiting Program !!!")

				workbook_sheet_list = openpyxl.load_workbook(ask).sheetnames

				for path_key in list(workbook_data.keys())[1:]:
					ask = input(f"Please Enter {path_key} Sheet Name => ")

					if (ask not in workbook_sheet_list):
						return print(f"Sheet Name {ask} is incorrect\nExiting Program !!!")
					workbook_data.update({path_key:ask})

				json_data_to_use = workbook_data
		
		except KeyboardInterrupt:
			print("\nThankyou For Using Give-My-Seat :)")
			exit()
		GiveMySeat.load_json(process_to_use, json_data_to_use)
		GiveMySeat.generate_seating(student_data, room_data)


if __name__ == "__main__":
	GiveMySeat.main()